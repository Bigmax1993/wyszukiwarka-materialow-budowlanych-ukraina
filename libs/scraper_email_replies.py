# -*- coding: utf-8 -*-
"""
Wspólny moduł: synchronizacja odpowiedzi Gmail (IMAP), wyciąganie cen (mail + PDF),
kolumny CRM w Excelu, czerwone wiersze gdy brak odpowiedzi po X godzinach.
"""
from __future__ import annotations

import email
import imaplib
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from email.header import decode_header
from email.utils import parseaddr, parsedate_to_datetime
from pathlib import Path
from typing import Any

import random
import time

from polish_text import (
    configure_utf8_environment,
    normalize_row_dict,
    normalize_unicode_text,
    sanitize_special_text,
)
from scraper_env import (
    ENV_GMAIL_APP_PASSWORD,
    ENV_GMAIL_SENDER_NAME,
    ENV_GMAIL_USER,
    get_env_value,
    get_gmail_app_password,
    get_gmail_user,
)
from mail_transport import get_imap_host, send_smtp_email

configure_utf8_environment()

# --- Konfiguracja domyślna ---
DEFAULT_NO_REPLY_HOURS = 3.0
DEFAULT_IMAP_FULL_SCAN_DAYS = 120
DEFAULT_REMINDER_MIN_HOURS = 3.0
REMINDER_1_HOURS_AFTER_SENT = 3.0
REMINDER_2_HOURS_AFTER_FIRST = 2.0
MAX_REMINDERS_PER_CONTACT = 2
REMINDER_SEND_DELAY_MIN = 22
REMINDER_SEND_DELAY_MAX = 58
# Kampania UA: przypomnienia co 3 dni (nie godzin)
UA_REMINDER_INTERVAL_DAYS = 3
UA_REMINDER_INTERVAL_HOURS = UA_REMINDER_INTERVAL_DAYS * 24.0
# UA: jedno przypomnienie, liczone od email_sent_at (pierwszego zapytania)
UA_MAX_REMINDERS_PER_CONTACT = 1
SEND_WINDOW_START = 8
SEND_WINDOW_END = 18

SIGNATURE_PL = (
    "Z poważaniem,\n\n"
    "Maksym Swinczak\n"
    "Kanbud Sp. z o.o.\n"
    "tel. +49 1522 3655 399"
)
# W stopce przypomnień DE pozostaje MFG (kampanie niemieckie)
SIGNATURE_DE = (
    "Mit freundlichen Grüßen\n"
    "Maksym Swinczak – MFG Modernerfliesenboden\n"
    "Tel.: +49 152 23655399\n"
    "Web: mfg-fliesen.de"
)
SIGNATURE_UK = (
    "З повагою,\n\n"
    "Свінчак Максим\n"
    "Tel.: +380977091141"
)
RED_FILL_RGB = "FFFFC7CE"
YELLOW_FILL_RGB = "FFFFFF99"

REPLY_EXPORT_COLUMNS = [
    "Status maila",
    "Wysłano",
    "Odpowiedź",
    "Status odpowiedzi",
    "Wymaga interwencji",
    "Odczytane (Twoja reakcja)",
    "Cena",
    "Waluta",
    "Opis",
    "Ceny (wszystkie)",
    "Źródło ceny",
    "Zadzwoń?",
]

ORANGE_FILL_RGB = "FFFFE699"

# Kotwice relacji (transport PL) — dopasowanie cen w tekście/OCR
TRANSPORT_RELATION_ANCHORS_PL = [
    ("rel_1", ("etoy", "1163", "szwajcar", " ch")),
    ("rel_2", ("saalfeld", "07318", "rewe", "niemiec", " de")),
    ("rel_3", ("opole", "odra", "cementownia")),
]

AUTO_REPLY_HINTS = (
    "automatic reply",
    "auto-reply",
    "autoreply",
    "out of office",
    "poza biurem",
    "nieobecno",
    "abwesenheit",
    "automatische antwort",
    "automatyczna odpowied",
    "urlaub",
    "vacation",
)

BOUNCE_HINTS = (
    "mail delivery failed",
    "undelivered",
    "undeliverable",
    "delivery status notification",
    "nie dostarczono",
    "nie została dostarczona",
    "nie zostala dostarczona",
    "mailbox unavailable",
    "user unknown",
    "unknown user",
    "użytkownik nieznany",
    "uzytkownik nieznany",
    "nie znaleziono adresu",
    "address not found",
    "recipient address rejected",
    "550 ",
    "553 ",
    "554 ",
)

BOUNCE_SENDER_LOCALPARTS = (
    "mailer-daemon",
    "mailerdaemon",
    "postmaster",
    "mail delivery subsystem",
)

EMAIL_IN_BOUNCE_RE = re.compile(
    r"[a-zA-Z0-9._%+\-']+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
    re.IGNORECASE,
)

QUESTION_HINTS_PL = (
    "proszę o",
    "prosze o",
    "czy ",
    "jaki nip",
    "jaki numer",
    "nie mamy w systemie",
    "proszę doprecyzować",
    "prosze doprecyzowac",
    "która firma",
    "ktora firma",
    "doprecyzow",
    "jakie dane",
)

QUESTION_HINTS_DE = (
    "bitte um",
    "können sie",
    "koennen sie",
    "welche",
    "ust-id",
    "steuernummer",
    "nicht im system",
    "rückfrage",
    "rueckfrage",
    "bitte teilen",
)

PRICE_PATTERNS = [
    re.compile(
        r"(?P<val>\d{1,3}(?:[\s\u00a0.]?\d{3})*(?:[.,]\d{1,2})?)\s*(?P<cur>zł|zl|PLN|EUR|€|CHF)\b",
        re.IGNORECASE,
    ),
    re.compile(
        r"(?P<cur>EUR|€|PLN|zł|zl|CHF)\s*(?P<val>\d{1,3}(?:[\s\u00a0.]?\d{3})*(?:[.,]\d{1,2})?)",
        re.IGNORECASE,
    ),
]

# Stopka firmowa / rejestrowa — nie jest wyceną transportu
FALSE_PRICE_SNIPPET_HINTS = (
    "kapitał zakładowy",
    "kapital zakladowy",
    "share capital",
    "registered capital",
    " w całości wpłacony",
    " w calosci wplacony",
    "krs ",
    "regon ",
    "nip ",
    "bdo ",
    "sąd rejonowy",
    "sad rejonowy",
)

POSITIVE_PRICE_SNIPPET_HINTS = (
    "stawka",
    "wycen",
    "cena",
    "oferta",
    "koszt",
    "eur/",
    "€/",
    "pln/",
    "za ton",
    "za transport",
    "relacja",
    "ftl",
    "all-in",
    "fracht",
    "preis",
    "angebot",
)


@dataclass
class ReplySyncConfig:
    cache_path: Path
    xlsx_path: Path
    lang: str = "pl"
    campaign_id: str = ""
    no_reply_hours: float = DEFAULT_NO_REPLY_HOURS
    relation_anchors: list[tuple[str, tuple[str, ...]]] = field(
        default_factory=lambda: list(TRANSPORT_RELATION_ANCHORS_PL)
    )
    imap_days_back: int = 14
    main_sheet_names: tuple[str, ...] = ("Kontakte", "Kontakty", "Baza firm")
    email_column: str = "E-mail"


UNKNOWN_COMPANY_LABEL_PL = "Nieznana firma"
UNKNOWN_COMPANY_LABEL_DE = "Unbekanntes Unternehmen"
UNKNOWN_COMPANY_LABELS = (UNKNOWN_COMPANY_LABEL_PL, UNKNOWN_COMPANY_LABEL_DE)
_UNKNOWN_LABELS_LOWER = {label.lower() for label in UNKNOWN_COMPANY_LABELS}

EXPORT_NAME_COLUMNS = (
    "Nazwa firmy",
    "Firma",
    "Nazwa Firmy",
    "Firmenname",
    "nazwa",
    "company_name_clean",
)


def is_unknown_company_name(name: str) -> bool:
    return (name or "").strip().lower() in _UNKNOWN_LABELS_LOWER


def export_row_company_name(row: dict) -> str:
    for key in EXPORT_NAME_COLUMNS:
        val = row.get(key)
        if val is not None and str(val).strip():
            return str(val).strip()
    return ""


def is_unknown_company_row(row: dict) -> bool:
    return is_unknown_company_name(export_row_company_name(row))


def should_skip_scraper_row(row: dict) -> bool:
    """Pomija puste wpisy „Nieznana firma” / „Unbekanntes Unternehmen” bez URL."""
    name = (row.get("company_name_clean") or row.get("nazwa") or "").strip()
    if not is_unknown_company_name(name):
        return False
    return not (row.get("url") or "").strip()


def filter_unknown_company_rows(rows: list[dict]) -> list[dict]:
    return [r for r in rows if not is_unknown_company_row(r)]


def purge_unknown_from_cache(cache: dict) -> int:
    """Usuwa wpisy „Nieznana firma” / „Unbekanntes Unternehmen” z cache."""
    removed = 0
    for section in ("contacts", "claude_row_enrichment", "gemini_row_enrichment"):
        bucket = cache.get(section)
        if not isinstance(bucket, dict):
            continue
        drop_keys = []
        for key, info in bucket.items():
            if not isinstance(info, dict):
                continue
            name = (
                info.get("company_name_clean")
                or info.get("company_name")
                or info.get("nazwa")
                or ""
            ).strip()
            key_low = str(key).lower()
            key_is_unknown = any(
                key_low.startswith(label.lower()) for label in UNKNOWN_COMPANY_LABELS
            )
            if key_is_unknown or is_unknown_company_name(name):
                drop_keys.append(key)
        for key in drop_keys:
            bucket.pop(key, None)
            removed += 1
    return removed


def normalize_email(addr: str) -> str:
    if not addr:
        return ""
    _, bare = parseaddr(addr)
    return (bare or addr).strip().lower()


def email_domain(addr: str) -> str:
    em = normalize_email(addr)
    if "@" in em:
        return em.split("@", 1)[1]
    return ""


# Domeny współdzielone — dopasowanie po domenie dawało jedną odpowiedź setkom firm (gmail.com itd.)
GENERIC_EMAIL_DOMAINS = frozenset(
    {
        "gmail.com",
        "googlemail.com",
        "wp.pl",
        "o2.pl",
        "op.pl",
        "interia.pl",
        "onet.pl",
        "poczta.onet.pl",
        "vp.pl",
        "outlook.com",
        "hotmail.com",
        "live.com",
        "yahoo.com",
        "yahoo.pl",
        "icloud.com",
        "protonmail.com",
        "proton.me",
        "gmx.de",
        "gmx.net",
        "web.de",
        "t-online.de",
        "aol.com",
        "mail.com",
        "spoko.pl",
        "autograf.pl",
        "go2.pl",
    }
)

REPLY_SYNC_FIELD_KEYS = (
    "reply_at",
    "reply_from",
    "reply_subject",
    "reply_status",
    "has_reply",
    "reply_body_snippet",
    "reply_description",
    "price_main",
    "price_currency",
    "prices_all",
    "price_source",
    "price_rel",
    "price_evidence",
    "needs_human_review",
    "extraction_method",
    "extraction_version",
    "reply_matched_by",
    "requires_intervention",
    "imap_uid",
)


def is_generic_email_domain(domain: str) -> bool:
    d = (domain or "").strip().lower()
    if not d:
        return True
    if d in GENERIC_EMAIL_DOMAINS:
        return True
    return any(d == g or d.endswith("." + g) for g in GENERIC_EMAIL_DOMAINS)


def mark_email_sent(
    contact: dict,
    subject: str,
    *,
    body: str = "",
    lang: str = "pl",
    campaign_id: str = "",
    relation_labels: list[str] | None = None,
) -> None:
    contact["email_sent_at"] = datetime.now().isoformat(timespec="seconds")
    contact["email_subject_sent"] = (subject or "").strip()
    if body and str(body).strip():
        contact["email_body_sent"] = str(body).strip()
    contact["email_lang"] = lang
    if campaign_id:
        contact["campaign_id"] = campaign_id
    if relation_labels:
        contact["relation_labels"] = relation_labels


def build_email_lookup(cache: dict) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    """Adres docelowy wysyłki (email_target) -> [place_url]; domena firmowa (nie gmail/wp) -> [place_url]."""
    by_email: dict[str, list[str]] = {}
    by_domain: dict[str, list[str]] = {}
    contacts = cache.get("contacts") or {}
    for place_url, info in contacts.items():
        if not isinstance(info, dict):
            continue
        target = normalize_email(info.get("email_target") or "")
        if not target:
            continue
        by_email.setdefault(target, [])
        if place_url not in by_email[target]:
            by_email[target].append(place_url)
        dom = email_domain(target)
        if dom and not is_generic_email_domain(dom):
            by_domain.setdefault(dom, [])
            if place_url not in by_domain[dom]:
                by_domain[dom].append(place_url)
    return by_email, by_domain


def _filter_urls_for_incoming_reply(cache: dict, urls: list[str], from_em: str) -> list[str]:
    """Tylko kontakty, na które faktycznie wysłano mail (email_target + email_sent_at)."""
    contacts = cache.get("contacts") or {}
    matched: list[str] = []
    for place_url in urls:
        info = contacts.get(place_url)
        if not isinstance(info, dict):
            continue
        if normalize_email(info.get("email_target") or "") != from_em:
            continue
        if not info.get("email_sent_at"):
            continue
        matched.append(place_url)
    return matched


def pick_best_place_url_for_reply(
    cache: dict, urls: list[str], msg_dt: datetime
) -> str | None:
    """Jeden wiersz na odpowiedź — najbliższa wysyłka przed datą maila."""
    if not urls:
        return None
    if len(urls) == 1:
        return urls[0]
    contacts = cache.get("contacts") or {}
    best_url: str | None = None
    best_sent: datetime | None = None
    for place_url in urls:
        info = contacts.get(place_url)
        if not isinstance(info, dict):
            continue
        raw = info.get("email_sent_at")
        if not raw:
            continue
        try:
            sent_at = datetime.fromisoformat(str(raw).replace("Z", ""))
        except ValueError:
            continue
        if sent_at > msg_dt + timedelta(minutes=5):
            continue
        if best_sent is None or sent_at > best_sent:
            best_sent = sent_at
            best_url = place_url
    return best_url or urls[0]


def resolve_place_urls(
    from_addr: str,
    by_email: dict,
    by_domain: dict,
    cache: dict | None = None,
) -> list[str]:
    em = normalize_email(from_addr)
    if not em:
        return []
    urls: list[str] = []
    if em in by_email:
        urls = list(by_email[em])
    else:
        dom = email_domain(em)
        if dom and not is_generic_email_domain(dom) and dom in by_domain:
            candidates = list(by_domain[dom])
            if len(candidates) == 1:
                urls = candidates
    if not urls:
        return []
    if cache is not None:
        return _filter_urls_for_incoming_reply(cache, urls, em)
    return urls


def clear_reply_fields(contact: dict) -> None:
    for key in REPLY_SYNC_FIELD_KEYS:
        contact.pop(key, None)
    contact["has_reply"] = False
    contact.pop("user_read_at", None)


def repair_misassigned_replies(cache: dict, logger: logging.Logger | None = None) -> int:
    """Usuwa odpowiedzi przypisane do złego kontaktu (np. fan-out po gmail.com)."""
    cleared = 0
    for _url, info in (cache.get("contacts") or {}).items():
        if not isinstance(info, dict):
            continue
        if not info.get("reply_at"):
            continue
        reply_from = normalize_email(info.get("reply_from") or "")
        target = normalize_email(info.get("email_target") or "")
        if not reply_from:
            continue
        if contact_is_bounced(info):
            continue
        if target and reply_from == target:
            continue
        clear_reply_fields(info)
        cleared += 1
    if cleared and logger:
        logger.info(f"Usunięto błędnie przypisane odpowiedzi: {cleared}")
    return cleared


def decode_mime_header(value: str | None) -> str:
    if not value:
        return ""
    parts = []
    for chunk, enc in decode_header(value):
        if isinstance(chunk, bytes):
            parts.append(chunk.decode(enc or "utf-8", errors="replace"))
        else:
            parts.append(str(chunk))
    return normalize_unicode_text("".join(parts).strip())


def html_to_text(html: str) -> str:
    if not html:
        return ""
    text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", html)
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</p>", "\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = unescape_basic(text)
    return " ".join(text.split())


def unescape_basic(s: str) -> str:
    return (
        s.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
    )


def get_message_body(msg: email.message.Message) -> str:
    plain_parts: list[str] = []
    html_parts: list[str] = []
    if msg.is_multipart():
        for part in msg.walk():
            ctype = (part.get_content_type() or "").lower()
            disp = str(part.get("Content-Disposition") or "").lower()
            if "attachment" in disp:
                continue
            payload = part.get_payload(decode=True)
            if payload is None:
                continue
            charset = part.get_content_charset() or "utf-8"
            try:
                text = payload.decode(charset, errors="replace")
            except Exception:
                text = payload.decode("utf-8", errors="replace")
            if ctype == "text/plain":
                plain_parts.append(text)
            elif ctype == "text/html":
                html_parts.append(text)
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or "utf-8"
            try:
                text = payload.decode(charset, errors="replace")
            except Exception:
                text = payload.decode("utf-8", errors="replace")
            if (msg.get_content_type() or "").lower() == "text/html":
                html_parts.append(text)
            else:
                plain_parts.append(text)
    if plain_parts:
        return normalize_unicode_text("\n".join(plain_parts).strip())
    if html_parts:
        return normalize_unicode_text(html_to_text("\n".join(html_parts)))
    return ""


def iter_pdf_attachments(msg: email.message.Message):
    for part in msg.walk():
        disp = str(part.get("Content-Disposition") or "").lower()
        fname = part.get_filename() or ""
        ctype = (part.get_content_type() or "").lower()
        if "attachment" not in disp and not fname.lower().endswith(".pdf"):
            if ctype != "application/pdf":
                continue
        if not (fname.lower().endswith(".pdf") or ctype == "application/pdf"):
            continue
        payload = part.get_payload(decode=True)
        if payload:
            yield fname or "attachment.pdf", payload


def extract_text_from_pdf_bytes(data: bytes, logger: logging.Logger) -> tuple[str, str]:
    """Zwraca (tekst, źródło: pdf_text|pdf_ocr|empty)."""
    text = _pdf_native_text(data)
    if len(text.strip()) >= 40:
        return text, "pdf_text"
    ocr_text = _pdf_ocr_text(data, logger)
    if len(ocr_text.strip()) >= 20:
        return ocr_text, "pdf_ocr"
    return text or ocr_text, "empty"


def _pdf_native_text(data: bytes) -> str:
    try:
        import fitz  # PyMuPDF  # pyright: ignore[reportMissingImports]

        doc = fitz.open(stream=data, filetype="pdf")
        parts = [page.get_text() for page in doc]
        doc.close()
        return "\n".join(parts)
    except Exception:
        pass
    try:
        from pypdf import PdfReader  # pyright: ignore[reportMissingImports]
        import io

        reader = PdfReader(io.BytesIO(data))
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        return "\n".join(parts)
    except Exception:
        return ""


def _pdf_ocr_text(data: bytes, logger: logging.Logger) -> str:
    try:
        import io

        import pytesseract  # pyright: ignore[reportMissingImports]
        from pdf2image import convert_from_bytes  # pyright: ignore[reportMissingImports]

        images = convert_from_bytes(data, dpi=300)
        lang = "pol+deu+eng"
        chunks = [pytesseract.image_to_string(img, lang=lang) for img in images]
        return "\n".join(chunks)
    except Exception as e:
        logger.info(f"OCR PDF pominięty (brak pdf2image/pytesseract?): {e}")
        return ""


def parse_message_date(msg: email.message.Message) -> datetime | None:
    raw = msg.get("Date")
    if not raw:
        return None
    try:
        dt = parsedate_to_datetime(raw)
        if dt.tzinfo:
            return dt.replace(tzinfo=None)
        return dt
    except Exception:
        return None


def classify_reply_text(text: str, lang: str = "pl") -> str:
    low = (text or "").lower()
    if any(h in low for h in BOUNCE_HINTS):
        return "bounce"
    if any(h in low for h in AUTO_REPLY_HINTS):
        return "auto_reply"
    prices = extract_price_candidates(text)
    if prices:
        return "replied_with_price"
    q_hints = list(QUESTION_HINTS_PL)
    q_hints.extend(QUESTION_HINTS_DE)
    if "?" in text and any(h in low for h in q_hints):
        return "replied_questions"
    if any(h in low for h in q_hints):
        return "replied_questions"
    if len(low.strip()) > 30:
        return "replied_no_price"
    return "replied_no_price"


def is_bounce_sender(from_addr: str) -> bool:
    em = normalize_email(from_addr)
    if not em:
        return False
    local = em.split("@", 1)[0]
    return any(h in local for h in BOUNCE_SENDER_LOCALPARTS)


def is_bounce_notification(from_addr: str, subject: str, body: str) -> bool:
    if is_bounce_sender(from_addr):
        return True
    combined = f"{subject or ''}\n{body or ''}".lower()
    return any(h in combined for h in BOUNCE_HINTS)


def extract_bounce_recipients(
    text: str,
    *,
    known_targets: set[str] | None = None,
) -> list[str]:
    """Adresy odbiorcy z treści DSN (bounce); preferuj znane email_target z cache."""
    skip_localparts = BOUNCE_SENDER_LOCALPARTS + (
        "noreply",
        "no-reply",
        "do-not-reply",
    )
    found: list[str] = []
    seen: set[str] = set()
    for match in EMAIL_IN_BOUNCE_RE.finditer(text or ""):
        em = normalize_email(match.group(0))
        if not em or em in seen:
            continue
        local = em.split("@", 1)[0]
        if any(h in local for h in skip_localparts):
            continue
        if em.endswith("@googlemail.com") and "mailer" in local:
            continue
        seen.add(em)
        found.append(em)
    if known_targets:
        matched = [em for em in found if em in known_targets]
        if matched:
            return matched
    return found


def contact_is_bounced(contact: dict) -> bool:
    if str(contact.get("email_status") or "").strip().lower() == "bounced":
        return True
    return str(contact.get("reply_status") or "").strip().lower() == "bounce"


def suppress_reminders_for_bounced_contact(
    contact: dict,
    *,
    reply_at: str = "",
    snippet: str = "",
) -> bool:
    """Trwale wyłącz przypomnienia po bounce (nieznany adres, 550 itd.)."""
    already = contact_is_bounced(contact) and contact.get("reminders_suppressed")
    if reply_at and not contact.get("reply_at"):
        contact["reply_at"] = reply_at
    if snippet and not (contact.get("reply_body_snippet") or "").strip():
        contact["reply_body_snippet"] = snippet[:2000]
    contact["reply_status"] = "bounce"
    contact["has_reply"] = False
    contact["reminders_suppressed"] = True
    contact["email_status"] = "bounced"
    contact["reminder_status"] = "skipped_bounce"
    contact["requires_intervention"] = False
    contact["intervention_status"] = ""
    return not already


def _parse_amount(val: str) -> float | None:
    s = (val or "").strip().replace("\u00a0", " ").replace(" ", "")
    if not s:
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = parts[0].replace(".", "") + "." + parts[1]
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(".", "") if s.count(".") > 1 else s
    try:
        v = float(s)
        if 5 <= v <= 9_999_999:
            return v
    except ValueError:
        pass
    return None


def _normalize_currency(cur: str) -> str:
    c = (cur or "").strip().upper()
    if c in ("€", "EURO"):
        return "EUR"
    if c in ("ZŁ", "ZL"):
        return "PLN"
    return c


def extract_price_candidates(text: str) -> list[dict]:
    found: list[dict] = []
    seen: set[tuple[float, str]] = set()
    body = text or ""
    for pat in PRICE_PATTERNS:
        for m in pat.finditer(body):
            if _is_false_price_context(body, m.start("val")):
                continue
            val = _parse_amount(m.group("val"))
            cur = _normalize_currency(m.group("cur"))
            if val is None or not cur:
                continue
            key = (round(val, 2), cur)
            if key in seen:
                continue
            seen.add(key)
            start = max(0, m.start() - 80)
            end = min(len(text), m.end() + 80)
            snippet = " ".join(text[start:end].split())
            rel_key = _match_relation(snippet.lower())
            found.append(
                {
                    "value": val,
                    "currency": cur,
                    "snippet": snippet,
                    "relation_key": rel_key,
                }
            )
    return found


def _match_relation(snippet_low: str) -> str:
    for rel_key, anchors in TRANSPORT_RELATION_ANCHORS_PL:
        if any(a in snippet_low for a in anchors):
            return rel_key
    return ""


def _is_false_price_context(text: str, match_start: int) -> bool:
    """Stopka rejestrowa tuż przed kwotą (np. kapitał zakładowy 603.060,00 zł)."""
    before = (text or "")[max(0, match_start - 45) : match_start].lower()
    return any(h in before for h in FALSE_PRICE_SNIPPET_HINTS)


def _price_candidate_score(p: dict) -> tuple[int, int, float]:
    """Wyższy = lepszy kandydat (nie największa liczba w stopce)."""
    sn = (p.get("snippet") or "").lower()
    score = 0
    if p.get("relation_key"):
        score += 100
    if any(h in sn for h in POSITIVE_PRICE_SNIPPET_HINTS):
        score += 50
    val = float(p.get("value") or 0)
    if 50 <= val <= 50_000:
        score += 20
    elif val > 100_000:
        score -= 80
    return (score, 1 if p.get("relation_key") else 0, -val)


def _pick_main_price(candidates: list[dict]) -> dict | None:
    if not candidates:
        return None
    return max(candidates, key=_price_candidate_score)


def merge_price_extractions(
    body_prices: list[dict], pdf_prices: list[dict], pdf_source: str
) -> dict[str, Any]:
    all_p = body_prices + pdf_prices
    if not all_p:
        return {
            "price_main": "",
            "price_currency": "",
            "prices_all": "",
            "reply_description": "",
            "price_source": "",
            "price_rel": {},
        }
    main = _pick_main_price(all_p)
    if not main:
        return {
            "price_main": "",
            "price_currency": "",
            "prices_all": "",
            "reply_description": "",
            "price_source": "",
            "price_rel": {},
        }
    rel_map: dict[str, str] = {}
    for p in all_p:
        rk = p.get("relation_key") or ""
        if rk and rk not in rel_map:
            rel_map[rk] = f"{p['value']:.2f} {p['currency']}"
    parts = [f"{p['value']:.2f} {p['currency']}" for p in all_p[:8]]
    source = "mail"
    if pdf_prices and body_prices:
        source = "mail+pdf"
    elif pdf_prices:
        source = pdf_source or "pdf"
    desc = main.get("snippet", "")[:400]
    return {
        "price_main": f"{main['value']:.2f}",
        "price_currency": main["currency"],
        "prices_all": "; ".join(parts),
        "reply_description": desc,
        "price_source": source,
        "price_rel": rel_map,
    }


def is_sent_status(status: str) -> bool:
    return str(status or "").strip().lower() == "sent"


def has_meaningful_reply(reply_status: str) -> bool:
    rs = (reply_status or "").strip().lower()
    if rs in ("auto_reply", "bounce", ""):
        return False
    return rs.startswith("replied")


def contact_has_inbound_reply(contact: dict) -> bool:
    """Wykryta wiadomość zwrotna od odbiorcy (IMAP / cache)."""
    if contact.get("reply_at"):
        return True
    if contact.get("reply_from"):
        return True
    if (contact.get("reply_body_snippet") or "").strip():
        return True
    if contact.get("has_reply"):
        return True
    return bool((contact.get("reply_status") or "").strip())


def contact_has_any_reply(contact: dict) -> bool:
    """Brak przypomnień — odpowiedź, bounce lub trwała blokada."""
    if contact.get("reminders_suppressed"):
        return True
    status = str(contact.get("email_status") or "").strip().lower()
    if status in ("replied", "bounced"):
        return True
    if contact_is_bounced(contact):
        return True
    return contact_has_inbound_reply(contact)


def suppress_reminders_for_replied_contact(contact: dict) -> bool:
    """
    Po wykryciu odpowiedzi — trwale wyłącz przypomnienia (także gdy odp. przyszła przed 3 dniami).
    Zwraca True jeśli kontakt został oznaczony po raz pierwszy.
    """
    if contact.get("reminders_suppressed") and str(
        contact.get("email_status") or ""
    ).strip().lower() in ("replied", "bounced"):
        return False
    if contact_is_bounced(contact):
        return suppress_reminders_for_bounced_contact(contact)
    if not contact_has_inbound_reply(contact):
        return False
    contact["reminders_suppressed"] = True
    status = str(contact.get("email_status") or "").strip().lower()
    if status in ("sent", "reminder_sent"):
        contact["email_status"] = "replied"
    if not (contact.get("reminder_status") or "").strip():
        contact["reminder_status"] = "skipped_has_reply"
    return True


def backfill_reminder_suppression_for_replies(
    cache: dict, logger: logging.Logger | None = None
) -> int:
    """Istniejące wpisy z odpowiedzią — bez kolejnych przypomnień."""
    updated = 0
    for info in (cache.get("contacts") or {}).values():
        if not isinstance(info, dict):
            continue
        if info.get("reminders_suppressed"):
            continue
        if suppress_reminders_for_replied_contact(info):
            updated += 1
    if logger and updated:
        logger.info(
            "backfill_reminder_suppression: %s kontakt(ów) z odpowiedzią — bez przypomnień",
            updated,
        )
    return updated


def requires_user_intervention(reply_status: str, body_text: str = "") -> bool:
    """Mail wymaga Twojej odpowiedzi (pytania, bounce, doprecyzowanie)."""
    rs = (reply_status or "").strip().lower()
    if rs in ("replied_questions", "bounce"):
        return True
    if rs == "replied_no_price":
        low = (body_text or "").lower()
        if "?" in (body_text or "") and any(h in low for h in QUESTION_HINTS_PL + QUESTION_HINTS_DE):
            return True
    return False


def is_user_marked_read(contact: dict) -> bool:
    return bool(contact.get("user_marked_read"))


def update_intervention_flags(contact: dict, reply_status: str, body_text: str) -> None:
    needs = requires_user_intervention(reply_status, body_text)
    contact["requires_intervention"] = needs
    if needs and not is_user_marked_read(contact):
        contact["user_marked_read"] = False
        contact["intervention_status"] = "nieodczytane"
    elif needs and is_user_marked_read(contact):
        contact["intervention_status"] = "odczytane"
    else:
        contact["intervention_status"] = ""


def compute_call_needed(contact: dict, no_reply_hours: float) -> bool:
    if not is_sent_status(contact.get("email_status", "")):
        return False
    if has_meaningful_reply(contact.get("reply_status", "")):
        return False
    if contact.get("has_reply"):
        return False
    sent_raw = contact.get("email_sent_at") or ""
    if not sent_raw:
        return False
    try:
        sent_at = datetime.fromisoformat(sent_raw.replace("Z", ""))
    except ValueError:
        return False
    now = datetime.now()
    if not (SEND_WINDOW_START <= now.hour < SEND_WINDOW_END):
        return False
    return (now - sent_at) >= timedelta(hours=float(no_reply_hours))


def reply_status_label(status: str, lang: str = "pl") -> str:
    m = {
        "replied_with_price": "Oferta",
        "replied_questions": "Pytania",
        "replied_no_price": "Bez ceny",
        "auto_reply": "Auto",
        "bounce": "Bounce",
    }
    if lang == "de":
        m = {
            "replied_with_price": "Angebot",
            "replied_questions": "Rückfrage",
            "replied_no_price": "Kein Preis",
            "auto_reply": "Auto",
            "bounce": "Bounce",
        }
    elif lang == "uk":
        m = {
            "replied_with_price": "Пропозиція",
            "replied_questions": "Питання",
            "replied_no_price": "Без ціни",
            "auto_reply": "Авто",
            "bounce": "Bounce",
        }
    return m.get((status or "").strip().lower(), status or "")


def export_columns_from_contact(contact: dict, lang: str = "pl") -> dict[str, str]:
    sent = (contact.get("email_sent_at") or "")[:19].replace("T", " ")
    reply_at = (contact.get("reply_at") or "")[:19].replace("T", " ")
    rs = (contact.get("reply_status") or "").strip()
    has_r = "Tak" if has_meaningful_reply(rs) or contact.get("has_reply") else "Nie"
    if lang == "de":
        has_r = "Ja" if has_r == "Tak" else "Nein"
    elif lang == "uk":
        has_r = "Так" if has_r == "Tak" else "Ні"
    call = contact.get("call_needed")
    if call is None:
        call = compute_call_needed(contact, DEFAULT_NO_REPLY_HOURS)
    call_l = "TAK" if call else ""
    rel_prices = contact.get("price_rel") or {}
    desc = (contact.get("reply_description") or "").strip()
    if isinstance(rel_prices, dict) and rel_prices:
        rel_txt = "; ".join(f"{k}: {v}" for k, v in rel_prices.items())
        if rel_txt:
            desc = (desc + " | " + rel_txt).strip(" |")
    needs_int = bool(contact.get("requires_intervention"))
    if lang == "de":
        int_lbl = "JA" if needs_int else ""
        read_lbl = (
            "Ja"
            if is_user_marked_read(contact)
            else ("Nein" if needs_int else "")
        )
    elif lang == "uk":
        int_lbl = "ТАК" if needs_int else ""
        read_lbl = (
            "Так"
            if is_user_marked_read(contact)
            else ("Ні" if needs_int else "")
        )
    else:
        int_lbl = "TAK" if needs_int else ""
        read_lbl = (
            "Tak"
            if is_user_marked_read(contact)
            else ("Nie" if needs_int else "")
        )
    row = {
        "Status maila": (contact.get("email_status") or "").strip(),
        "Wysłano": sent,
        "Odpowiedź": has_r,
        "Status odpowiedzi": reply_status_label(rs, lang) if rs else "",
        "Wymaga interwencji": int_lbl,
        "Odczytane (Twoja reakcja)": read_lbl,
        "Cena": (contact.get("price_main") or "").strip(),
        "Waluta": (contact.get("price_currency") or "").strip(),
        "Opis": sanitize_special_text(desc)[:500],
        "Ceny (wszystkie)": (contact.get("prices_all") or "").strip(),
        "Źródło ceny": (contact.get("price_source") or "").strip(),
        "Zadzwoń?": call_l,
    }
    anchors = contact.get("relation_labels") or []
    if not anchors and rel_prices:
        anchors = ["Rel. 1", "Rel. 2", "Rel. 3"]
    for i, rk in enumerate(("rel_1", "rel_2", "rel_3"), start=1):
        col = f"Cena rel. {i}"
        if isinstance(rel_prices, dict):
            row[col] = rel_prices.get(rk, "")
        else:
            row[col] = ""
    return row


def get_contact_for_email(cache: dict, email_addr: str) -> dict | None:
    em = normalize_email(email_addr)
    if not em:
        return None
    by_email, by_domain = build_email_lookup(cache)
    urls = resolve_place_urls(em, by_email, by_domain, cache)
    if not urls:
        return None
    return (cache.get("contacts") or {}).get(urls[0])


def find_contact_by_export_email(cache: dict, email_col_value: str) -> dict | None:
    return get_contact_for_email(cache, email_col_value)


def apply_reply_to_contact(
    contact: dict,
    *,
    reply_at: datetime,
    reply_from: str,
    reply_subject: str,
    reply_status: str,
    body_text: str,
    pdf_text: str,
    pdf_source: str,
    lang: str,
    gemini_cache: dict | None = None,
    matched_by: str = "email",
) -> None:
    from email_reply_intelligence import analyze_incoming_reply, strip_quoted_reply

    contact["reply_at"] = reply_at.isoformat(timespec="seconds")
    contact["reply_from"] = normalize_email(reply_from)
    contact["reply_subject"] = (reply_subject or "")[:500]
    contact["reply_matched_by"] = matched_by
    body_clean = strip_quoted_reply(body_text)
    contact["reply_body_snippet"] = (body_clean or body_text or "")[:2000]

    forced_bounce = (reply_status or "").strip().lower() == "bounce"
    if forced_bounce:
        analyzed: dict = {}
    else:
        analyzed = analyze_incoming_reply(
            contact=contact,
            from_email=normalize_email(reply_from),
            subject=reply_subject or "",
            body_text=body_text,
            pdf_text=pdf_text,
            pdf_source=pdf_source,
            lang=lang,
            claude_cache=gemini_cache,
        )
        reply_status = analyzed.get("reply_status") or reply_status
    contact["reply_status"] = reply_status
    contact["has_reply"] = has_meaningful_reply(reply_status)

    for key in (
        "price_main",
        "price_currency",
        "prices_all",
        "price_source",
        "price_rel",
        "reply_description",
        "price_evidence",
        "needs_human_review",
        "extraction_method",
        "extraction_version",
    ):
        if key in analyzed:
            contact[key] = analyzed[key]
    if not contact.get("price_source"):
        contact["price_source"] = "mail" if pdf_text else "mail"
    if reply_status == "replied_questions" and not contact.get("price_main"):
        q = analyzed.get("questions_text") or ""
        contact["reply_description"] = (q or body_clean or "")[:500]
    contact["call_needed"] = compute_call_needed(contact, DEFAULT_NO_REPLY_HOURS)
    update_intervention_flags(contact, reply_status, body_clean or body_text)
    if reply_status == "bounce":
        suppress_reminders_for_bounced_contact(
            contact,
            reply_at=contact.get("reply_at") or "",
            snippet=contact.get("reply_body_snippet") or "",
        )
    else:
        suppress_reminders_for_replied_contact(contact)


def fetch_imap_messages(
    logger: logging.Logger,
    since: datetime,
    *,
    mark_intervention_unread: bool = True,
) -> list[tuple[bytes, email.message.Message]]:
    user = get_gmail_user()
    password = get_gmail_app_password()
    imap_host = get_imap_host()
    if not (user and password):
        raise RuntimeError("Brak MAIL_USER / MAIL_PASSWORD (lub GMAIL_*) w zmiennych środowiskowych.")
    if not imap_host:
        raise RuntimeError("Brak IMAP_HOST w zmiennych środowiskowych.")

    messages: list[tuple[bytes, email.message.Message]] = []
    try:
        from email_journal import log_imap_scan_start

        log_imap_scan_start(since)
    except Exception:
        pass
    logger.info(f"IMAP: skan INBOX od {since.strftime('%Y-%m-%d')}…")
    since_str = since.strftime("%d-%b-%Y")
    mail = imaplib.IMAP4_SSL(imap_host)
    try:
        mail.login(user, password)
        mail.select("INBOX")
        typ, data = mail.search(None, f'(SINCE "{since_str}")')
        if typ != "OK" or not data or not data[0]:
            logger.info("IMAP: brak nowych wiadomości w zakresie.")
            return []
        for num in data[0].split():
            typ2, msg_data = mail.fetch(num, "(BODY.PEEK[])")
            if typ2 != "OK" or not msg_data:
                continue
            for part in msg_data:
                if isinstance(part, tuple) and len(part) > 1:
                    msg = email.message_from_bytes(part[1])
                    messages.append((num, msg))
        logger.info(
            f"IMAP: pobrano {len(messages)} wiadomości "
            f"(BODY.PEEK — skan nie oznacza maili jako przeczytane w Gmail)."
        )
        try:
            from email_journal import log_imap_scan_end

            log_imap_scan_end(len(messages))
        except Exception:
            pass
    finally:
        try:
            mail.logout()
        except Exception:
            pass
    return messages


def normalize_imap_items(
    messages: list,
) -> list[tuple[bytes, email.message.Message]]:
    items: list[tuple[bytes, email.message.Message]] = []
    for item in messages or []:
        if isinstance(item, tuple) and len(item) >= 2:
            items.append((item[0], item[1]))
        else:
            items.append((b"", item))
    return items


def mark_imap_messages_unread(logger: logging.Logger, uids: list[bytes]) -> int:
    """Oznacza wiadomości w Gmail jako nieprzeczytane (usuwa flagę \\Seen)."""
    if not uids:
        return 0
    user = get_gmail_user()
    password = get_gmail_app_password()
    imap_host = get_imap_host()
    if not (user and password and imap_host):
        logger.warning("Brak konfiguracji IMAP do oznaczenia wiadomości jako nieprzeczytane.")
        return 0
    mail = imaplib.IMAP4_SSL(imap_host)
    marked = 0
    try:
        mail.login(user, password)
        mail.select("INBOX")
        for uid in uids:
            try:
                typ, _ = mail.store(uid, "-FLAGS", "\\Seen")
                if typ == "OK":
                    marked += 1
            except Exception as e:
                logger.debug(f"IMAP store {uid}: {e}")
        logger.info(f"IMAP: oznaczono jako nieprzeczytane: {marked}/{len(uids)}")
    finally:
        try:
            mail.logout()
        except Exception:
            pass
    return marked


def _imap_since_for_cache(
    cache: dict, config: ReplySyncConfig, full_inbox_scan: bool
) -> datetime:
    if full_inbox_scan:
        return datetime.now() - timedelta(days=config.imap_days_back)
    since_raw = cache.get("email_sync_last_at")
    if since_raw:
        try:
            return datetime.fromisoformat(str(since_raw).replace("Z", ""))
        except ValueError:
            pass
    return datetime.now() - timedelta(days=config.imap_days_back)


def sync_replies_from_messages(
    messages: list,
    cache: dict,
    config: ReplySyncConfig,
    logger: logging.Logger,
    since: datetime,
    *,
    force_refresh: bool = False,
) -> tuple[int, list[bytes]]:
    repair_misassigned_replies(cache, logger)
    by_email, by_domain = build_email_lookup(cache)
    contacts = cache.setdefault("contacts", {})
    claude_cache = cache.setdefault("claude_reply_extractions", {})
    if not claude_cache and isinstance(cache.get("gemini_reply_extractions"), dict):
        claude_cache.update(cache.get("gemini_reply_extractions") or {})
    updated = 0
    uids_mark_unread: list[bytes] = []
    our_email = normalize_email(get_gmail_user())
    campaign = config.campaign_id or ""
    known_targets = set(by_email.keys())

    for imap_uid, msg in normalize_imap_items(messages):
        msg_dt = parse_message_date(msg) or datetime.now()
        if msg_dt < since - timedelta(days=1):
            continue
        from_hdr = decode_mime_header(msg.get("From"))
        from_em = normalize_email(from_hdr)
        if not from_em or from_em == our_email:
            continue

        subject = decode_mime_header(msg.get("Subject"))
        body = get_message_body(msg)

        if is_bounce_notification(from_em, subject, body):
            recipients = extract_bounce_recipients(
                f"{subject}\n{body}", known_targets=known_targets
            )
            for recip in recipients:
                place_urls = resolve_place_urls(
                    recip, by_email, by_domain, cache
                )
                if not place_urls:
                    continue
                place_url = pick_best_place_url_for_reply(cache, place_urls, msg_dt)
                if not place_url:
                    continue
                contact = contacts.setdefault(place_url, {})
                sent_raw = contact.get("email_sent_at")
                if sent_raw:
                    try:
                        sent_at = datetime.fromisoformat(sent_raw.replace("Z", ""))
                        if msg_dt < sent_at - timedelta(minutes=5):
                            continue
                    except ValueError:
                        pass
                if contact_is_bounced(contact):
                    continue
                apply_reply_to_contact(
                    contact,
                    reply_at=msg_dt,
                    reply_from=from_em,
                    reply_subject=subject,
                    reply_status="bounce",
                    body_text=body,
                    pdf_text="",
                    pdf_source="",
                    lang=config.lang,
                    gemini_cache=claude_cache,
                    matched_by="bounce",
                )
                company = (
                    contact.get("company_name_clean")
                    or contact.get("company_name")
                    or recip
                )
                logger.info(
                    "Bounce dla %s → %s — temat: %s",
                    recip,
                    company,
                    (subject or "")[:50],
                )
                updated += 1
            continue

        place_urls = resolve_place_urls(from_em, by_email, by_domain, cache)
        if not place_urls:
            continue

        pdf_text = ""
        pdf_source = ""
        for fname, payload in iter_pdf_attachments(msg):
            logger.info(f"PDF z maila: {fname}")
            pdf_text, pdf_source = extract_text_from_pdf_bytes(payload, logger)
            if pdf_text.strip():
                break

        place_url = pick_best_place_url_for_reply(cache, place_urls, msg_dt)
        if not place_url:
            continue
        contact = contacts.setdefault(place_url, {})
        sent_raw = contact.get("email_sent_at")
        if sent_raw:
            try:
                sent_at = datetime.fromisoformat(sent_raw.replace("Z", ""))
                if msg_dt < sent_at - timedelta(minutes=5):
                    continue
            except ValueError:
                pass
        if not force_refresh:
            prev_reply = contact.get("reply_at")
            if prev_reply:
                try:
                    if msg_dt <= datetime.fromisoformat(prev_reply.replace("Z", "")):
                        continue
                except ValueError:
                    pass

        apply_reply_to_contact(
            contact,
            reply_at=msg_dt,
            reply_from=from_em,
            reply_subject=subject,
            reply_status="",
            body_text=body,
            pdf_text=pdf_text,
            pdf_source=pdf_source,
            lang=config.lang,
            gemini_cache=claude_cache,
            matched_by="email",
        )
        company = (
            contact.get("company_name_clean")
            or contact.get("company_name")
            or from_em
        )
        reply_status = contact.get("reply_status") or ""
        logger.info(
            f"Mail od {from_em} → {company} ({reply_status}) — temat: {(subject or '')[:50]}"
        )
        try:
            from email_journal import log_mail_read

            log_mail_read(
                from_em,
                subject,
                matched=True,
                reply_status=reply_status,
                company=str(company),
                campaign=campaign,
            )
        except Exception:
            pass
        if imap_uid and contact.get("requires_intervention") and not is_user_marked_read(
            contact
        ):
            contact["imap_uid"] = (
                imap_uid.decode() if isinstance(imap_uid, bytes) else str(imap_uid)
            )
            uids_mark_unread.append(imap_uid)
        updated += 1

    for info in contacts.values():
        if isinstance(info, dict):
            if is_sent_status(info.get("email_status", "")):
                info["call_needed"] = compute_call_needed(info, config.no_reply_hours)
            if info.get("requires_intervention") and not is_user_marked_read(info):
                if not info.get("intervention_status"):
                    info["intervention_status"] = "nieodczytane"

    cache["email_sync_last_at"] = datetime.now().isoformat(timespec="seconds")
    try:
        from email_journal import log_sync_summary

        log_sync_summary(updated, campaign)
    except Exception:
        pass
    return updated, uids_mark_unread


def sync_replies_to_cache(
    cache: dict,
    config: ReplySyncConfig,
    logger: logging.Logger,
    *,
    full_inbox_scan: bool = False,
    prefetched_messages: list | None = None,
    force_refresh: bool = False,
    mark_gmail_unread: bool = True,
) -> int:
    since = _imap_since_for_cache(cache, config, full_inbox_scan)
    if prefetched_messages is None:
        prefetched_messages = fetch_imap_messages(logger, since)
    updated, uids = sync_replies_from_messages(
        prefetched_messages, cache, config, logger, since, force_refresh=force_refresh
    )
    if mark_gmail_unread and uids:
        mark_imap_messages_unread(logger, uids)
    return updated


def get_all_campaign_presets(base_dir: Path) -> dict[str, dict]:
    """Wszystkie kampanie PL + DE + CH (cache + xlsx)."""
    from project_config import switzerland_wyniki_dir

    base = Path(base_dir)
    ch_dir = switzerland_wyniki_dir()
    presets = {
        "szczecin": {
            "cache": base / "Wyniki_Szczecin" / "szczecin_piasek_cache.json",
            "xlsx": base / "Wyniki_Szczecin" / "szczecin_piasek_kontakte.xlsx",
            "lang": "pl",
            "campaign_id": "sand_szczecin",
            "sheets": ("Kontakte", "Wojewodztwa"),
            "main_sheets": ("Kontakte",),
        },
        "warszawa": {
            "cache": base / "Wyniki_Warszawa" / "warszawa_piasek_cache.json",
            "xlsx": base / "Wyniki_Warszawa" / "warszawa_piasek_kontakte.xlsx",
            "lang": "pl",
            "campaign_id": "sand_warszawa",
            "sheets": ("Kontakte", "Wojewodztwa"),
            "main_sheets": ("Kontakte",),
        },
        "transport": {
            "cache": base / "Wyniki_Transport_PL" / "poland_transport_cache.json",
            "xlsx": base / "Wyniki_Transport_PL" / "poland_transport_firmy.xlsx",
            "lang": "pl",
            "campaign_id": "transport_pl",
            "sheets": ("Kontakty", "Baza firm", "Wojewodztwa"),
            "main_sheets": ("Kontakty", "Baza firm"),
        },
        "saalfeld_sand": {
            "cache": base / "Wyniki_Saalfeld" / "saalfeld_sand_kies_cache.json",
            "xlsx": base / "Wyniki_Saalfeld" / "saalfeld_sand_kies_kontakte.xlsx",
            "lang": "de",
            "campaign_id": "sand_saalfeld_de",
            "sheets": ("Kontakte", "Bundeslaender"),
            "main_sheets": ("Kontakte",),
        },
        "saalfeld_beton": {
            "cache": base / "Wyniki_Saalfeld_Beton" / "saalfeld_betonwerk_cache.json",
            "xlsx": base / "Wyniki_Saalfeld_Beton" / "saalfeld_betonwerk_kontakte.xlsx",
            "lang": "de",
            "campaign_id": "beton_saalfeld_de",
            "sheets": ("Kontakte", "Bundeslaender"),
            "main_sheets": ("Kontakte",),
        },
        "saalfeld_baucontainer": {
            "cache": base / "Wyniki_Saalfeld_Baucontainer" / "saalfeld_baucontainer_cache.json",
            "xlsx": base / "Wyniki_Saalfeld_Baucontainer" / "saalfeld_baucontainer_kontakte.xlsx",
            "lang": "de",
            "campaign_id": "baucontainer_saalfeld_de",
            "sheets": ("Kontakte", "Bundeslaender"),
            "main_sheets": ("Kontakte",),
        },
        "saalfeld_radlader": {
            "cache": base / "Wyniki_Saalfeld_Radlader" / "saalfeld_radlader_cache.json",
            "xlsx": base / "Wyniki_Saalfeld_Radlader" / "saalfeld_radlader_kontakte.xlsx",
            "lang": "de",
            "campaign_id": "radlader_saalfeld_de",
            "sheets": ("Kontakte", "Bundeslaender"),
            "main_sheets": ("Kontakte",),
        },
    }
    ch_presets = {
        "switzerland_sand": {
            "cache": ch_dir / "switzerland_sand_gravel_cache.json",
            "xlsx": ch_dir / "switzerland_sand_gravel_contacts.xlsx",
            "lang": "de",
            "campaign_id": "sand_switzerland_ch",
            "sheets": ("Kontakte", "Kanton"),
            "main_sheets": ("Kontakte",),
        },
        "switzerland_hotels": {
            "cache": ch_dir / "switzerland_hotels_cache.json",
            "xlsx": ch_dir / "switzerland_hotels_contacts.xlsx",
            "lang": "de",
            "campaign_id": "hotels_switzerland_ch",
            "sheets": ("Kontakte",),
            "main_sheets": ("Kontakte",),
        },
    }
    for key, preset in ch_presets.items():
        if preset["cache"].is_file() or preset["xlsx"].is_file():
            presets[key] = preset
    return presets


def preset_to_config(preset: dict, no_reply_hours: float, imap_days: int) -> ReplySyncConfig:
    return ReplySyncConfig(
        cache_path=Path(preset["cache"]),
        xlsx_path=Path(preset["xlsx"]),
        lang=preset.get("lang", "pl"),
        campaign_id=preset.get("campaign_id", ""),
        no_reply_hours=no_reply_hours,
        imap_days_back=imap_days,
        main_sheet_names=tuple(preset.get("main_sheets", preset.get("sheets", ("Kontakte",)))),
    )


def sanitize_sender_name(name: str) -> str:
    return " ".join((name or "").replace("\n", " ").split()).strip()


def send_email_gmail(
    to_email: str,
    subject: str,
    body: str,
    logger: logging.Logger,
    *,
    mail_type: str = "wiadomość",
    campaign: str = "",
) -> tuple[bool, str]:
    sender_name = sanitize_sender_name(get_env_value(ENV_GMAIL_SENDER_NAME))
    if sender_name and not get_env_value("MAIL_SENDER_NAME"):
        try:
            import os

            os.environ.setdefault("MAIL_SENDER_NAME", sender_name)
        except Exception:
            pass
    ok, info = send_smtp_email(
        to_email,
        sanitize_special_text(subject),
        sanitize_special_text(body),
        logger,
        mail_type=mail_type,
        campaign=campaign,
    )
    return ok, ("reminder_sent" if ok else info)


def _parse_sent_at(contact: dict) -> datetime | None:
    raw = contact.get("email_sent_at") or ""
    if not raw:
        return None
    try:
        return datetime.fromisoformat(str(raw).replace("Z", ""))
    except ValueError:
        return None


def _parse_dt(raw: str | None) -> datetime | None:
    if not raw:
        return None
    try:
        return datetime.fromisoformat(str(raw).replace("Z", ""))
    except ValueError:
        return None


def had_reply_within_days_of_sent(contact: dict, days: float) -> bool:
    """
    True jeśli odbiorca odpowiedział przed upływem ``days`` od email_sent_at.
    Gdy brak reply_at, ale jest ślad odpowiedzi — traktuj jako blokujące (bezpiecznie).
    """
    if not contact_has_inbound_reply(contact):
        return False
    sent_at = _parse_sent_at(contact)
    reply_at = _parse_dt(contact.get("reply_at"))
    if sent_at and reply_at:
        return reply_at < sent_at + timedelta(days=float(days))
    return True


def get_ua_pending_reminder_number(
    contact: dict,
    *,
    min_days: float | None = None,
) -> int | None:
    """
    UA: jedno przypomnienie dopiero po min_days od pierwszego maila (email_sent_at).
    Pomija kontakty z odpowiedzią w tym oknie (i każdą inną odpowiedź).
    """
    if min_days is None:
        min_days = UA_REMINDER_INTERVAL_DAYS
    min_hours = float(min_days) * 24.0

    if contact_has_any_reply(contact) or had_reply_within_days_of_sent(contact, min_days):
        suppress_reminders_for_replied_contact(contact)
        return None

    return get_pending_reminder_number(
        contact,
        first_after_hours=min_hours,
        second_after_hours=min_hours,
        max_reminders=UA_MAX_REMINDERS_PER_CONTACT,
    )


def reminder_count(contact: dict, *, max_reminders: int | None = None) -> int:
    """Ile przypomnień już wysłano."""
    cap = MAX_REMINDERS_PER_CONTACT if max_reminders is None else int(max_reminders)
    raw = contact.get("reminder_count")
    if raw is not None and str(raw).strip() != "":
        try:
            return max(0, min(int(raw), cap))
        except (TypeError, ValueError):
            pass
    if cap >= 2 and contact.get("reminder_2_sent_at"):
        return 2
    if contact.get("reminder_sent_at"):
        return 1
    return 0


def get_pending_reminder_number(
    contact: dict,
    *,
    first_after_hours: float | None = None,
    second_after_hours: float | None = None,
    max_reminders: int | None = None,
) -> int | None:
    """
    1 = pierwsze przypomnienie po first_after_hours od zapytania,
    2 = drugie po second_after_hours od pierwszego przypomnienia (kampanie PL/DE).
    """
    if first_after_hours is None:
        first_after_hours = REMINDER_1_HOURS_AFTER_SENT
    if second_after_hours is None:
        second_after_hours = REMINDER_2_HOURS_AFTER_FIRST
    cap = MAX_REMINDERS_PER_CONTACT if max_reminders is None else int(max_reminders)

    if contact.get("reminders_suppressed"):
        return None
    status = str(contact.get("email_status") or "").strip().lower()
    if status in ("replied", "bounced"):
        return None
    if contact_is_bounced(contact):
        return None

    if contact_has_any_reply(contact):
        return None
    if not (contact.get("email_target") or "").strip():
        return None

    count = reminder_count(contact, max_reminders=cap)
    if count >= cap:
        return None

    now = datetime.now()

    if count == 0:
        if str(contact.get("email_status") or "").strip().lower() != "sent":
            return None
        sent_at = _parse_sent_at(contact)
        if not sent_at:
            return None
        if now - sent_at < timedelta(hours=float(first_after_hours)):
            return None
        return 1

    if count == 1 and cap >= 2:
        first_at = _parse_dt(contact.get("reminder_sent_at"))
        if not first_at:
            return None
        if now - first_at < timedelta(hours=float(second_after_hours)):
            return None
        return 2

    return None


def needs_reminder(
    contact: dict,
    min_hours: float | None = None,
    *,
    second_after_hours: float | None = None,
    max_reminders: int | None = None,
) -> bool:
    """True jeśli należy wysłać kolejne przypomnienie (1. lub 2.)."""
    if min_hours is None:
        min_hours = DEFAULT_REMINDER_MIN_HOURS
    if second_after_hours is None:
        second_after_hours = min_hours
    pending = get_pending_reminder_number(
        contact,
        first_after_hours=min_hours,
        second_after_hours=second_after_hours,
        max_reminders=max_reminders,
    )
    return pending is not None


def ua_needs_reminder(contact: dict, min_days: float | None = None) -> bool:
    """UA: jedno przypomnienie po min_days od pierwszego maila, bez odpowiedzi w tym oknie."""
    return get_ua_pending_reminder_number(contact, min_days=min_days) is not None


def mark_reminder_sent(contact: dict, which: int) -> None:
    """Zapisuje wysłanie przypomnienia 1 lub 2 (nie więcej niż 2 łącznie)."""
    which = int(which)
    if which not in (1, 2):
        raise ValueError("which must be 1 or 2")
    now = datetime.now().isoformat(timespec="seconds")
    if which == 1:
        contact["reminder_sent_at"] = now
        contact["reminder_count"] = 1
        contact["reminder_status"] = "sent_1"
        contact["email_status"] = "reminder_sent"
    else:
        contact["reminder_2_sent_at"] = now
        contact["reminder_count"] = 2
        contact["reminder_status"] = "sent_2"
        contact["email_status"] = "reminders_complete"


SIGNATURE_START_PATTERNS = (
    r"(?i)^z\s+poważaniem",
    r"(?i)^z\s+powazaniem",
    r"(?i)^pozdrawiam\b",
    r"(?i)^mit\s+freundlichen\s+gr",
    r"(?i)^freundliche\s+gr",
    r"(?i)^beste\s+gr",
    r"(?i)^viele\s+gr",
)


def get_original_email_body(contact: dict) -> str:
    """Treść pierwszego wysłanego maila (do stopki i cytatu)."""
    for key in ("email_body_sent", "email_body"):
        raw = (contact.get(key) or "").strip()
        if raw:
            return raw
    return ""


def split_body_and_signature(body: str) -> tuple[str, str]:
    """Rozdziela treść główną i stopkę (od linii z pozdrowieniem)."""
    text = (body or "").strip()
    if not text:
        return "", ""
    lines = text.splitlines()
    sig_idx: int | None = None
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped:
            continue
        for pat in SIGNATURE_START_PATTERNS:
            if re.match(pat, stripped):
                sig_idx = i
                break
        if sig_idx is not None:
            break
    if sig_idx is None:
        return text, ""
    main = "\n".join(lines[:sig_idx]).strip()
    signature = "\n".join(lines[sig_idx:]).strip()
    return main, signature


def _format_sent_datetime(sent_raw: str) -> str:
    if not sent_raw:
        return ""
    try:
        dt = datetime.fromisoformat(str(sent_raw).replace("Z", ""))
        return dt.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return sent_raw[:16].replace("T", " ")


def _reply_subject(original_subject: str, lang: str, company: str) -> str:
    subj = (original_subject or "").strip()
    while subj:
        low = subj.lower()
        if low.startswith("re:"):
            subj = subj[3:].strip()
        elif low.startswith("odp:"):
            subj = subj[4:].strip()
        elif low.startswith("aw:"):
            subj = subj[3:].strip()
        else:
            break
    if subj:
        return f"Re: {subj}"
    if lang == "de":
        return f"Re: Preisanfrage – {company}"
    if lang == "uk":
        return f"Re: запит щодо пропозиції – {company}"
    return f"Re: zapytanie ofertowe – {company}"


def _is_mfg_signature_text(text: str) -> bool:
    low = (text or "").lower()
    return any(
        x in low
        for x in (
            "mfg modernerfliesenboden",
            "mfg-fliesen",
            "mfg fliesen",
            "– mfg",
            "- mfg",
        )
    )


def normalize_signature_for_pl(signature: str) -> str:
    """Przypomnienia do firm PL — stopka Kanbud, nie MFG."""
    sig = (signature or "").strip()
    if not sig or _is_mfg_signature_text(sig):
        return SIGNATURE_PL
    if "kanbud" in sig.lower():
        return sig
    # Zamiana linii firmowej MFG na Kanbud, zachowaj resztę jeśli sensowna
    lines = sig.splitlines()
    out: list[str] = []
    for line in lines:
        if _is_mfg_signature_text(line):
            if "Kanbud" not in "\n".join(out):
                out.extend(["Maksym Swinczak", "Kanbud Sp. z o.o.", "tel. +49 1522 3655 399"])
            continue
        if "mfg-fliesen" in line.lower() or line.strip().lower().startswith("www:"):
            continue
        out.append(line)
    merged = "\n".join(out).strip()
    return merged if merged and "kanbud" in merged.lower() else SIGNATURE_PL


def _reminder_from_line(lang: str) -> str:
    sender = get_gmail_user() or get_env_value(ENV_GMAIL_USER) or ""
    if lang == "de":
        sender_name = sanitize_sender_name(get_env_value(ENV_GMAIL_SENDER_NAME))
        if sender_name and sender:
            return f"{sender_name} <{sender}>"
        return sender or sender_name or ""
    if lang == "uk":
        sender_name = sanitize_sender_name(
            get_env_value(ENV_GMAIL_SENDER_NAME) or get_env_value("MAIL_SENDER_NAME")
        )
        if not sender_name:
            sender_name = "Свінчак Максим"
        if sender:
            return f"{sender_name} <{sender}>"
        return sender_name
    # PL: zawsze Kanbud w nagłówku cytatu (nie nazwa nadawcy z Gmail/MFG)
    if sender:
        return f"Maksym Swinczak, Kanbud Sp. z o.o. <{sender}>"
    return "Maksym Swinczak, Kanbud Sp. z o.o."


def format_quoted_previous_email(
    original_body: str,
    original_subject: str,
    sent_at_raw: str,
    lang: str,
) -> str:
    """Blok historii poprzedniego maila (jak w odpowiedzi w wątku)."""
    if not (original_body or "").strip():
        return ""
    from_line = _reminder_from_line(lang)
    sent_fmt = _format_sent_datetime(sent_at_raw)

    if lang == "de":
        header = "--- Ursprüngliche Nachricht ---\n"
        if sent_fmt:
            header += f"Datum: {sent_fmt}\n"
        header += f"Betreff: {original_subject or '(ohne Betreff)'}\n"
        if from_line:
            header += f"Von: {from_line}\n"
    elif lang == "uk":
        header = "--- Попереднє повідомлення ---\n"
        if sent_fmt:
            header += f"Дата: {sent_fmt}\n"
        header += f"Тема: {original_subject or '(без теми)'}\n"
        if from_line:
            header += f"Від: {from_line}\n"
    else:
        header = "--- Poprzednia wiadomość ---\n"
        if sent_fmt:
            header += f"Data: {sent_fmt}\n"
        header += f"Temat: {original_subject or '(brak tematu)'}\n"
        if from_line:
            header += f"Od: {from_line}\n"

    quoted_body = original_body.strip()
    return f"{header}\n{quoted_body}"


def _static_reminder_intro_uk(
    sent_date: str, *, reminder_number: int = 1
) -> str:
    date_bit = f" від {sent_date}" if sent_date else ""
    if reminder_number >= 2:
        return (
            "Доброго дня,\n\n"
            f"Пишу щодо нашого запиту{date_bit} — "
            "на жаль, відповіді ще не отримали. "
            "Чи могли б Ви повернутися з короткою інформацією або орієнтовною ціною?\n\n"
            "Буду вдячний за будь-який зворотний зв'язок."
        )
    return (
        "Доброго дня,\n\n"
        f"Хотів би делікатно повернутися до нашого листа{date_bit} "
        "щодо будматеріалів. Можливо, він загубився серед інших запитів.\n\n"
        "Якщо є можливість, прошу коротко відповісти або надіслати прайс. "
        "Заздалегідь дякую."
    )


def normalize_signature_for_uk(signature: str) -> str:
    sig = (signature or "").strip()
    if not sig:
        return SIGNATURE_UK
    low = sig.lower()
    if any(x in low for x in ("+49", "0049", "mfg", "gmbh", "kanbud")):
        return SIGNATURE_UK
    if "380" in sig or "свінчак" in low or "swinczak" in low:
        return sig
    return SIGNATURE_UK


def build_reminder_email(
    contact: dict,
    lang: str,
    *,
    reminder_number: int = 1,
    logger: logging.Logger | None = None,
    cache: dict | None = None,
    place_url: str = "",
) -> tuple[str, str]:
    company = (
        contact.get("company_name_clean")
        or contact.get("company_name")
        or contact.get("nazwa")
        or "Firma"
    ).strip()
    orig_subj = (contact.get("email_subject_sent") or contact.get("email_subject") or "").strip()
    orig_body = get_original_email_body(contact)
    sent_raw = contact.get("email_sent_at") or ""
    sent_date = sent_raw[:10] if sent_raw else ""

    _main, signature = split_body_and_signature(orig_body)
    if lang == "de":
        if not signature:
            signature = SIGNATURE_DE
    elif lang == "uk":
        signature = normalize_signature_for_uk(signature)
    else:
        signature = normalize_signature_for_pl(signature)

    subject = _reply_subject(orig_subj, lang, company)

    intro = ""
    if lang == "uk":
        try:
            from ua_claude_reminder_email import claude_generate_reminder_intro_uk

            intro = claude_generate_reminder_intro_uk(
                contact,
                logger,
                cache,
                reminder_number=reminder_number,
                cache_key=place_url or str(contact.get("email_target") or ""),
            ) or ""
        except Exception:
            intro = ""
    if not intro:
        if reminder_number >= 2:
            if lang == "de":
                intro = (
                    f"Guten Tag,\n\n"
                    f"ich möchte unsere Anfrage"
                    f"{f' vom {sent_date}' if sent_date else ''} "
                    f"noch einmal kurz ansprechen — bisher haben wir leider keine "
                    f"Rückmeldung erhalten. Wären Sie so freundlich, uns zeitnah zu antworten?"
                )
            elif lang == "uk":
                intro = _static_reminder_intro_uk(sent_date, reminder_number=2)
            else:
                intro = (
                    f"Dzień dobry,\n\n"
                    f"ponownie przypominam o naszym zapytaniu ofertowym"
                    f"{f' z dnia {sent_date}' if sent_date else ''} — "
                    f"niestety nie otrzymaliśmy jeszcze odpowiedzi. "
                    f"Będę wdzięczny za krótką informację lub wycenę."
                )
        elif lang == "de":
            intro = (
                f"Guten Tag,\n\n"
                f"ich erlaube mir, unsere Anfrage"
                f"{f' vom {sent_date}' if sent_date else ''} "
                f"freundlich in Erinnerung zu rufen. Könnten Sie uns bitte kurz "
                f"rückmelden oder ein unverbindliches Angebot zusenden?"
            )
        elif lang == "uk":
            intro = _static_reminder_intro_uk(sent_date, reminder_number=1)
        else:
            intro = (
                f"Dzień dobry,\n\n"
                f"uprzejmie przypominam o naszym zapytaniu ofertowym"
                f"{f' z dnia {sent_date}' if sent_date else ''}. "
                f"Będę wdzięczny za krótką informację zwrotną lub wycenę."
            )

    quoted = format_quoted_previous_email(orig_body, orig_subj, sent_raw, lang)
    parts = [intro, signature]
    if quoted:
        parts.append(quoted)
    body = "\n\n".join(p for p in parts if p.strip())
    return subject, body


BILINGUAL_REMINDER_SEPARATOR = "\n\n──────────\n\n"


def backfill_email_sent_metadata(cache: dict, logger: logging.Logger | None = None) -> int:
    """Uzupełnia brakujące pola po wysłaniu zapytania (stare wpisy cache)."""
    updated = 0
    contacts = cache.get("contacts") or {}
    for _url, info in contacts.items():
        if not isinstance(info, dict):
            continue
        changed = False
        if not (info.get("email_subject_sent") or "").strip():
            subj = (info.get("email_subject") or "").strip()
            if subj:
                info["email_subject_sent"] = subj
                changed = True
        if not (info.get("email_body_sent") or "").strip():
            body = (info.get("email_body") or "").strip()
            if body:
                info["email_body_sent"] = body
                changed = True
        status = str(info.get("email_status") or "").strip().lower()
        has_sent_at = bool((info.get("email_sent_at") or "").strip())
        has_body = bool(
            (info.get("email_body_sent") or info.get("email_body") or "").strip()
        )
        if has_sent_at and has_body and status not in (
            "sent",
            "reminder_sent",
            "reminders_complete",
        ):
            if not status.startswith(("duplicate_", "deferred_", "suppressed_")):
                info["email_status"] = "sent"
                changed = True
        if changed:
            updated += 1
    if logger and updated:
        logger.info(
            "backfill_email_sent_metadata: zaktualizowano %s kontakt(ów)", updated
        )
    return updated


def _reminder_intro_fr(contact: dict, *, reminder_number: int) -> str:
    sent_raw = contact.get("email_sent_at") or ""
    sent_date = sent_raw[:10] if sent_raw else ""
    date_bit = f" du {sent_date}" if sent_date else ""
    if reminder_number >= 2:
        return (
            f"Bonjour,\n\n"
            f"je me permets de relancer notre demande{date_bit} — "
            f"nous n'avons pas encore reçu de retour. "
            f"Pourriez-vous nous répondre prochainement ?"
        )
    return (
        f"Bonjour,\n\n"
        f"je me permets de relancer notre demande{date_bit}. "
        f"Pourriez-vous nous répondre ou nous envoyer une offre sans engagement ?"
    )


def _build_bilingual_reminder_email_de_fr(
    contact: dict, *, reminder_number: int = 1
) -> tuple[str, str]:
    orig_subj = (
        contact.get("email_subject_sent") or contact.get("email_subject") or ""
    ).strip()
    orig_body = get_original_email_body(contact)
    sent_raw = contact.get("email_sent_at") or ""
    subject, body_de = build_reminder_email(contact, "de", reminder_number=reminder_number)
    intro_fr = _reminder_intro_fr(contact, reminder_number=reminder_number)
    quoted = format_quoted_previous_email(orig_body, orig_subj, sent_raw, "de")
    main_de, signature = split_body_and_signature(body_de)
    parts = [
        main_de,
        BILINGUAL_REMINDER_SEPARATOR,
        f"Madame, Monsieur,\n\n{intro_fr}",
        signature,
    ]
    if quoted:
        parts.append(quoted)
    body = "\n\n".join(p for p in parts if p and str(p).strip())
    return subject, body


def build_reminder_email_for_preset(
    contact: dict,
    preset: dict,
    *,
    reminder_number: int = 1,
    logger: logging.Logger | None = None,
    cache: dict | None = None,
    place_url: str = "",
) -> tuple[str, str]:
    lang = str(preset.get("lang") or "pl").strip().lower()
    if preset.get("bilingual_reminder_de_fr"):
        return _build_bilingual_reminder_email_de_fr(
            contact, reminder_number=reminder_number
        )
    return build_reminder_email(
        contact,
        lang,
        reminder_number=reminder_number,
        logger=logger,
        cache=cache,
        place_url=place_url,
    )


def _latest_inbound_from_target(
    contact: dict,
    messages: list,
    our_email: str,
) -> tuple[datetime, email.message.Message] | None:
    """Najnowsza wiadomość od email_target po dacie wysłania zapytania."""
    target = normalize_email(contact.get("email_target") or "")
    if not target:
        return None
    sent_at = _parse_sent_at(contact)
    if not sent_at:
        return None
    best_dt: datetime | None = None
    best_msg: email.message.Message | None = None
    for _uid, msg in normalize_imap_items(messages):
        from_em = normalize_email(decode_mime_header(msg.get("From")))
        if not from_em or from_em != target or from_em == our_email:
            continue
        msg_dt = parse_message_date(msg) or datetime.now()
        if msg_dt < sent_at - timedelta(minutes=5):
            continue
        if best_dt is None or msg_dt > best_dt:
            best_dt = msg_dt
            best_msg = msg
    if best_dt is None or best_msg is None:
        return None
    return best_dt, best_msg


def _latest_bounce_for_target(
    contact: dict,
    messages: list,
) -> tuple[datetime, email.message.Message] | None:
    """Najnowszy DSN (mailer-daemon) wskazujący na email_target kontaktu."""
    target = normalize_email(contact.get("email_target") or "")
    if not target:
        return None
    sent_at = _parse_sent_at(contact)
    if not sent_at:
        return None
    best_dt: datetime | None = None
    best_msg: email.message.Message | None = None
    for _uid, msg in normalize_imap_items(messages):
        from_em = normalize_email(decode_mime_header(msg.get("From")))
        subject = decode_mime_header(msg.get("Subject"))
        body = get_message_body(msg)
        if not is_bounce_notification(from_em, subject, body):
            continue
        recipients = extract_bounce_recipients(f"{subject}\n{body}")
        if target not in recipients:
            continue
        msg_dt = parse_message_date(msg) or datetime.now()
        if msg_dt < sent_at - timedelta(minutes=5):
            continue
        if best_dt is None or msg_dt > best_dt:
            best_dt = msg_dt
            best_msg = msg
    if best_dt is None or best_msg is None:
        return None
    return best_dt, best_msg


def verify_contact_reply_from_imap(
    contact: dict,
    config: ReplySyncConfig,
    messages: list,
    logger: logging.Logger | None = None,
    *,
    cache: dict | None = None,
) -> bool:
    """Sprawdza skrzynkę dla tego samego adresata (email_target). Zwraca True jeśli jest odpowiedź lub bounce."""
    our_email = normalize_email(get_gmail_user())
    found = _latest_bounce_for_target(contact, messages)
    if not found:
        found = _latest_inbound_from_target(contact, messages, our_email)
    if not found:
        return contact_has_any_reply(contact)
    msg_dt, msg = found
    prev_reply = contact.get("reply_at")
    if prev_reply:
        try:
            prev_dt = datetime.fromisoformat(str(prev_reply).replace("Z", ""))
            if msg_dt <= prev_dt:
                return contact_has_any_reply(contact)
        except ValueError:
            pass
    subject = decode_mime_header(msg.get("Subject"))
    body = get_message_body(msg)
    pdf_text = ""
    pdf_source = ""
    for fname, payload in iter_pdf_attachments(msg):
        if logger:
            logger.info(f"Weryfikacja PDF ({contact.get('email_target')}): {fname}")
        pdf_text, pdf_source = extract_text_from_pdf_bytes(payload, logger)
        if pdf_text.strip():
            break
    claude_cache = None
    if cache is not None:
        claude_cache = cache.setdefault("claude_reply_extractions", {})
        if not claude_cache and isinstance(cache.get("gemini_reply_extractions"), dict):
            claude_cache.update(cache.get("gemini_reply_extractions") or {})
    apply_reply_to_contact(
        contact,
        reply_at=msg_dt,
        reply_from=normalize_email(decode_mime_header(msg.get("From"))),
        reply_subject=subject,
        reply_status="",
        body_text=body,
        pdf_text=pdf_text,
        pdf_source=pdf_source,
        lang=config.lang,
        gemini_cache=claude_cache,
        matched_by="email",
    )
    if logger:
        logger.info(
            f"Weryfikacja IMAP: odpowiedź od {contact.get('email_target')} "
            f"({contact.get('reply_status')})"
        )
    return True


def verify_sent_contacts_from_imap(
    cache: dict,
    config: ReplySyncConfig,
    messages: list,
    logger: logging.Logger | None = None,
) -> int:
    """Dla kontaktów ze statusem sent/reminder_sent — weryfikacja odpowiedzi od email_target."""
    updated = 0
    for _url, info in (cache.get("contacts") or {}).items():
        if not isinstance(info, dict):
            continue
        status = str(info.get("email_status") or "").strip().lower()
        if status not in ("sent", "reminder_sent"):
            continue
        if not (info.get("email_target") or "").strip():
            continue
        if contact_is_bounced(info):
            continue
        had = contact_has_any_reply(info)
        verify_contact_reply_from_imap(info, config, messages, logger, cache=cache)
        if contact_has_any_reply(info):
            suppress_reminders_for_replied_contact(info)
            if not had:
                updated += 1
    return updated


def iter_ua_reminder_candidates(
    cache: dict,
    min_days: float | None = None,
    *,
    messages: list | None = None,
    config: ReplySyncConfig | None = None,
    logger: logging.Logger | None = None,
) -> list[tuple[str, dict, str]]:
    """Kandydaci UA: jedno przypomnienie po min_days od email_sent_at."""
    if min_days is None:
        min_days = UA_REMINDER_INTERVAL_DAYS
    if messages is not None and config is not None:
        verify_sent_contacts_from_imap(cache, config, messages, logger)
    out: list[tuple[str, dict, str]] = []
    for place_url, info in (cache.get("contacts") or {}).items():
        if not isinstance(info, dict):
            continue
        target = (info.get("email_target") or "").strip()
        if not target:
            continue
        if ua_needs_reminder(info, min_days=min_days):
            out.append((place_url, info, target))
    return out


def iter_reminder_candidates(
    cache: dict,
    min_hours: float | None = None,
    *,
    second_after_hours: float | None = None,
    max_reminders: int | None = None,
    messages: list | None = None,
    config: ReplySyncConfig | None = None,
    logger: logging.Logger | None = None,
) -> list[tuple[str, dict, str]]:
    """[(place_url, contact, email_target), ...] — kampanie PL/DE (do 2 przypomnień)."""
    if min_hours is None:
        min_hours = DEFAULT_REMINDER_MIN_HOURS
    if second_after_hours is None:
        second_after_hours = min_hours
    if messages is not None and config is not None:
        verify_sent_contacts_from_imap(cache, config, messages, logger)
    out: list[tuple[str, dict, str]] = []
    for place_url, info in (cache.get("contacts") or {}).items():
        if not isinstance(info, dict):
            continue
        target = (info.get("email_target") or "").strip()
        if not target:
            continue
        if needs_reminder(
            info,
            min_hours,
            second_after_hours=second_after_hours,
            max_reminders=max_reminders,
        ):
            out.append((place_url, info, target))
    return out


def sleep_between_reminder_sends(logger: logging.Logger, to_email: str) -> None:
    delay = round(random.uniform(REMINDER_SEND_DELAY_MIN, REMINDER_SEND_DELAY_MAX), 1)
    logger.info(f"Pauza {delay}s przed kolejnym mailem (po {to_email})")
    time.sleep(delay)


def load_cache(path: Path) -> dict:
    if not path.exists():
        return {"contacts": {}}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_cache(path: Path, cache: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def write_excel_with_reply_styles(
    path: Path,
    sheets: dict[str, list[dict]],
    cache: dict,
    config: ReplySyncConfig,
    logger: logging.Logger,
) -> None:
    import pandas as pd  # pyright: ignore[reportMissingImports]
    from openpyxl.styles import PatternFill  # pyright: ignore[reportMissingImports]

    path.parent.mkdir(parents=True, exist_ok=True)
    red_fill = PatternFill(start_color=RED_FILL_RGB, end_color=RED_FILL_RGB, fill_type="solid")
    yellow_fill = PatternFill(
        start_color=YELLOW_FILL_RGB, end_color=YELLOW_FILL_RGB, fill_type="solid"
    )
    orange_fill = PatternFill(
        start_color=ORANGE_FILL_RGB, end_color=ORANGE_FILL_RGB, fill_type="solid"
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, rows in sheets.items():
            rows = filter_unknown_company_rows(rows or [])
            if not rows:
                pd.DataFrame().to_excel(writer, index=False, sheet_name=sheet_name)
                continue
            enriched = []
            for row in rows:
                r = normalize_row_dict(dict(row))
                em = (r.get(config.email_column) or r.get("E-mail") or "").strip()
                contact = find_contact_by_export_email(cache, em) if em else None
                if contact:
                    extra = export_columns_from_contact(contact, config.lang)
                    for k, v in extra.items():
                        if k not in r or not str(r.get(k, "")).strip():
                            r[k] = v
                        elif k in REPLY_EXPORT_COLUMNS:
                            r[k] = v
                    r["_call_needed"] = compute_call_needed(contact, config.no_reply_hours)
                    r["_intervention_unread"] = bool(
                        contact.get("requires_intervention")
                    ) and not is_user_marked_read(contact)
                    r["_reply_questions"] = (
                        contact.get("reply_status") == "replied_questions"
                    )
                else:
                    r["_call_needed"] = False
                    r["_intervention_unread"] = False
                    r["_reply_questions"] = False
                enriched.append(r)
            df = pd.DataFrame(enriched)
            meta_cols = [c for c in df.columns if c.startswith("_")]
            df_export = df.drop(columns=meta_cols, errors="ignore")
            df_export.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            call_flags = df["_call_needed"].tolist() if "_call_needed" in df.columns else []
            int_flags = (
                df["_intervention_unread"].tolist()
                if "_intervention_unread" in df.columns
                else []
            )
            q_flags = df["_reply_questions"].tolist() if "_reply_questions" in df.columns else []
            for row_idx, (call, intervention, quest) in enumerate(
                zip(call_flags, int_flags, q_flags), start=2
            ):
                fill = None
                if call:
                    fill = red_fill
                elif intervention:
                    fill = orange_fill
                elif quest:
                    fill = yellow_fill
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

    logger.info(f"Zapisano Excel ze stylami: {path}")


def merge_export_row(base: dict, cache: dict, email_value: str, lang: str = "pl") -> dict:
    row = dict(base)
    contact = find_contact_by_export_email(cache, email_value)
    if contact:
        row.update(export_columns_from_contact(contact, lang))
    return row
